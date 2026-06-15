"""
excel_reader.py — reads the HarvestBridge Excel workbook dynamically.
Row numbers verified against actual workbook structure.
"""
import openpyxl

FIN_ROWS = {
    "loan_fees":                   7,
    "saas_revenue":                8,
    "insurance_commissions":       9,
    "other_income":                10,
    "total_revenue":               11,
    "tech_infrastructure":         13,
    "credit_loss_provision":       14,
    "third_party_data":            15,
    "gross_profit":                17,
    "sales_marketing":             19,
    "general_admin":               20,
    "rd":                          21,
    "ebitda":                      23,
    "depreciation_amortisation":   24,
    "ebit":                        25,
    "interest_expense":            27,
    "fx_loss":                     28,
    "income_tax":                  29,
    "net_income":                  30,
    "gross_margin":                33,
    "ebitda_margin":               34,
}

KPI_ROWS = {
    "cash_runway":               6,
    "fte_headcount":             7,
    "female_fte":                8,
    "female_board_pct":          9,
    "female_exec_pct":           10,
    "female_midmgmt_pct":        11,
    "loans_disbursed":           12,
    "active_insurance_policies": 13,
    "claims_paid":               14,
    "active_saas_users":         15,
    "distributor_partners":      16,
}


def _to_float(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def read_excel(filepath: str) -> dict:
    wb = openpyxl.load_workbook(filepath, data_only=True)

    # ── Financial Performance ─────────────────────────────────────
    ws = wb["Financial Performance"]

    # Quarter labels in row 5, columns B onwards (index 1+)
    header = list(ws.iter_rows(min_row=5, max_row=5, values_only=True))[0]
    fin_cols = []
    for i, cell in enumerate(header):
        if cell and isinstance(cell, str):
            label = cell.strip().split("\n")[0].strip()
            # Only completed quarters — starts with Q, no FY, no Proj, no vs
            if (label.startswith("Q")
                    and "FY" not in label
                    and "vs" not in label
                    and "Proj" not in cell):  # check raw cell for Proj
                fin_cols.append((i, label))

    fin_data = {}
    quarters = []
    for col_idx, label in fin_cols:
        row_data = {}
        for metric, row_num in FIN_ROWS.items():
            cells = list(ws.iter_rows(min_row=row_num, max_row=row_num,
                                      values_only=True))[0]
            row_data[metric] = _to_float(cells[col_idx])
        if row_data.get("total_revenue"):
            fin_data[label] = row_data
            quarters.append(label)

    # ── KPI Dashboard ─────────────────────────────────────────────
    ws_kpi = wb["KPI Dashboard"]

    # Labels in row 5, data columns start at index 2 (C)
    kpi_header = list(ws_kpi.iter_rows(min_row=5, max_row=5, values_only=True))[0]
    kpi_cols = []
    for i, cell in enumerate(kpi_header):
        if cell and isinstance(cell, str):
            label = cell.strip().rstrip(" →").strip()
            if label.startswith("Q"):
                kpi_cols.append((i, label))

    kpi_data = {}
    for col_idx, label in kpi_cols:
        row_data = {}
        for kpi, row_num in KPI_ROWS.items():
            cells = list(ws_kpi.iter_rows(min_row=row_num, max_row=row_num,
                                          values_only=True))[0]
            val = cells[col_idx]
            if val is not None:
                row_data[kpi] = val
        if row_data:
            kpi_data[label] = row_data

    prior_label   = quarters[-1] if quarters else None
    prior_quarter = fin_data.get(prior_label, {})

    return {
        "quarters":            quarters,
        "financials":          fin_data,
        "kpis":                kpi_data,
        "prior_quarter_label": prior_label,
        "prior_quarter":       prior_quarter,
    }
